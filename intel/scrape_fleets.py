#!/usr/bin/env python3
"""
Oxley Fleet Intelligence — TX + LA statewide fleet scraper (single file, config-driven).

Pipeline (each layer is independently runnable / gated):

  1. FMCSA spine (FREE) ....... Company Census (az4n-8mr2) + SMS Census (kjg3-diqy)
                                + per-carrier inspection OOS aggregate (fx4q-ay7w).
  2. Screens (FREE) ........... national blocklist, passenger/govt/school exclusion,
                                private-ownership SCORE 0-100 (soft), public hard-exclude.
  3. Verticals + tire-burn .... keyword classification -> Severe / Heavy / General.
  4. Self-service score ....... FMCSA-only 0-100 (power/driver ratio, vehicle-OOS rate
                                vs mileage, ownership of tractors/trailers). RUNS_OWN_SHOP tag.
  5. Geocode + residential .... free Census batch geocoder; commercial vs residential/PO-box.
  6. Apify deep-harvest ....... GATED + COST-CAPPED. Maps + contact/email + Indeed jobs.
  7. Output -> Google Drive ... one xlsx per metro/region, sheets split by vertical +
                                RUNS OWN SHOP + REVIEW tabs. SETX = its own deep file.
  8. Renewal (--diff) ......... SETX-only re-pull, diff vs last run, harvest only NEW.

FIRST-RUN CONTRACT: default mode runs ONLY layers 1-5 (all free), prints counts and the
Apify cost estimate, and STOPS. The paid Apify harvest (layer 6) never runs without an
explicit --harvest AND (--confirm or interactive y/n), and always aborts above --max-cost.

Secrets come from ENV ONLY (never hardcode):
  APIFY_TOKEN                     Apify runs (layer 6)
  MAPS_API_KEY                    optional Google geocode fallback
  GDRIVE_SERVICE_ACCOUNT_FILE     path to service-account json  (or)
  GDRIVE_SERVICE_ACCOUNT_JSON     inline service-account json
  SOCRATA_APP_TOKEN               optional, raises Socrata rate limits
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone

import requests

# =============================================================================
# CONFIG  (edit here)
# =============================================================================

STATES = ["TX", "LA"]                 # states to pull (physical location)

# --- truck-size floors: OFF. All quantities included. ---
POWER_UNITS_MIN = 0                   # 0 == no floor
DRIVERS_MIN = 0                       # 0 == no floor

# --- Apify deep-harvest gating ---
HARVEST_TOP_N = 150                   # top private-scored fleets statewide to harvest
MAX_COST_USD = 25.0                   # hard cap; harvest aborts if estimate exceeds this
HARVEST_INCLUDE_SETX = True           # always harvest every SETX carrier (in addition to top-N)
INDEED_MAX_POSTINGS_PER_FLEET = 5     # cap Indeed results per fleet (cost control)

# --- Google Drive ---
DRIVE_FOLDER_NAME = "Oxley Fleet Intel"
OWNER_EMAIL = "moxley@oxleytireinc.com"   # folder is shared to this account as writer

# --- SETX (Beaumont / Southeast Texas) ZIP list --------------------------------
# Jefferson, Orange, Hardin, Jasper, Newton, Tyler, Chambers, Liberty counties.
SETX_ZIPS = {
    # Beaumont
    "77701", "77702", "77703", "77704", "77705", "77706", "77707", "77708",
    "77709", "77710", "77713", "77720", "77725", "77726",
    # Port Arthur / Groves / Nederland / Port Neches
    "77640", "77641", "77642", "77643", "77619", "77627", "77651",
    # Orange / Bridge City / Vidor / West Orange / Mauriceville
    "77630", "77631", "77632", "77611", "77662", "77670", "77664",
    # Beaumont-area small towns
    "77705", "77713", "77659", "77585",
    # Hardin County — Silsbee, Kountze, Lumberton, Sour Lake, Saratoga
    "77656", "77657", "77625", "77659", "77585",
    # Jasper / Newton / Tyler / Kirbyville / Woodville / Buna / Call
    "75951", "75956", "75966", "75979", "77664", "77612", "77615",
    # Chambers — Anahuac, Winnie, Mont Belvieu, Hamshire, Stowell
    "77514", "77665", "77580", "77597", "77661",
    # Liberty — Liberty, Dayton, Cleveland, Hardin, Devers, Hull
    "77575", "77535", "77327", "77561", "77538", "77564",
}

# --- Metro/region bucketing by ZIP3 prefix (SETX list above takes precedence) ---
# Anything TX not matched -> "rest-of-TX". Any LA carrier -> "Louisiana".
REGION_ZIP3 = {
    # Houston metro (excludes SETX zips, which are pulled out first)
    "770": "Houston", "772": "Houston", "773": "Houston",
    "774": "Houston", "775": "Houston", "771": "Houston",
    # DFW
    "750": "DFW", "751": "DFW", "752": "DFW", "753": "DFW", "754": "DFW",
    "760": "DFW", "761": "DFW", "762": "DFW",
    # Austin
    "786": "Austin", "787": "Austin", "789": "Austin",
}
REGION_ORDER = ["SETX", "Houston", "DFW", "Austin", "rest-of-TX", "Louisiana"]

# --- National blocklist (case-insensitive substring on legal_name / dba_name) ---
NATIONAL_BLOCKLIST = [
    "swift", "werner", "j.b. hunt", "jb hunt", "j b hunt", "schneider",
    "knight-swift", "knight transportation", "us xpress", "u.s. xpress",
    "prime inc", "prime, inc", "landstar", "covenant", "heartland express",
    "cr england", "c.r. england", "estes", "old dominion", "saia",
    "xpo", "rxo", "arcbest", "abf freight", "tforce", "ryder", "penske",
    "united parcel", "ups inc", "ups ground", "fedex", "federal express",
    "amazon relay", "amazon logistics", "walmart transport", "wal-mart transport",
    "sysco", "mclane", "halliburton", "schlumberger", "slb ", "baker hughes",
    "waste management", "wm corporate", "republic services", "clean harbors",
]

# --- Corporate-services registered agents (NOT an individual owner signal) ---
CORPORATE_AGENTS = [
    "ct corporation", "c t corporation", "registered agents inc",
    "registered agent inc", "cogency global", "corporation service company",
    "csc ", "national registered agents", "incorp services", "legalzoom",
    "northwest registered agent", "harbor compliance", "united states corporation",
    "vcorp", "capitol services", "corporate creations", "paracorp",
]

# --- Public-company hard-exclude markers (belt-and-suspenders w/ blocklist) ---
PUBLIC_MARKERS = [
    " nyse", " nasdaq", "(nyse", "(nasdaq", " plc", " s.a.", " ag ",
    "holdings inc", "industries inc", "international inc",
]

# --- Passenger / government / school exclusion (name keywords; SMS flags primary) --
PASSENGER_GOVT_NAME = [
    "school district", " isd", "independent school", "county of", "city of",
    "town of", "parish of", "state of", "university", "college", "church",
    "ministries", "transit authority", "metro transit", "regional transit",
    "charter bus", "tour", "tours", "limousine", "limo ", "shuttle",
    "taxi", "ambulance", " ems", "medical transport", "paratransit",
    "head start", "boys and girls", "ymca", "senior center", "nursing",
    "coach lines", "trailways", "greyhound",
]

# --- Verticals & tire-burn tier (checked against legal_name + dba_name) ----------
# tier: Severe (oilfield/aggregate/heavy) > Heavy (waste/vac/crane) > General
SEVERE_KEYWORDS = [
    "oilfield", "oil field", "oilwell", "frac", "fracking", "sand", "proppant",
    "aggregate", "aggregates", "dump", "concrete", "ready mix", "readymix",
    "redi-mix", "heavy haul", "heavyhaul", "heavy-haul", "logging", "log ",
    "logs", "timber", "gravel", "rock", "dirt", "crude", "roustabout",
    "water transfer", "winch", "dozer", "excavat", "quarry", "asphalt",
    "cement", "mining", "mine ", "drilling", "rig ", "pipe", "pipeline",
    "haul", "hauling", "materials", "stone", "lime", "clay", "caliche",
    "roll off", "rolloff", "roll-off", "flatbed", "lowboy", "lowbed",
]
HEAVY_KEYWORDS = [
    "waste", "refuse", "garbage", "trash", "vacuum", "vac ", "vac truck",
    "septic", "grease", "crane", "demolition", "demo ", "scrap", "recycl",
    "portable toilet", "porta", "tank wash", "environmental", "disposal",
    "roustabout", "sweeper", "hydro", "jetting",
]
# Vertical label (for sheet grouping) derived from the same keyword hits.
VERTICAL_RULES = [
    ("Oilfield / Energy", ["oilfield", "oil field", "frac", "crude", "roustabout",
                            "drilling", "rig ", "pipeline", "well", "proppant",
                            "water transfer", "winch", "oilwell"]),
    ("Aggregate / Construction", ["aggregate", "sand", "gravel", "rock", "dirt",
                                   "concrete", "ready mix", "readymix", "cement",
                                   "asphalt", "dump", "materials", "stone", "lime",
                                   "caliche", "clay", "quarry", "excavat", "dozer",
                                   "construction", "paving", "grading"]),
    ("Logging / Timber", ["logging", "log ", "logs", "timber", "lumber", "forest",
                          "pulpwood", "chip"]),
    ("Waste / Environmental", ["waste", "refuse", "garbage", "trash", "recycl",
                               "disposal", "septic", "vacuum", "vac ", "grease",
                               "environmental", "roll off", "rolloff", "hydro",
                               "jetting", "sweeper"]),
    ("Heavy Haul / Specialized", ["heavy haul", "heavyhaul", "lowboy", "lowbed",
                                  "crane", "oversize", "rigging", "winch", "wrecker",
                                  "towing", "tow "]),
    ("Agriculture", ["farm", "cattle", "livestock", "grain", "hay", "poultry",
                     "dairy", "ag ", "agri", "harvest", "feed", "ranch"]),
    ("Logistics / General Freight", ["logistics", "freight", "transport", "trucking",
                                     "carrier", "express", "cartage", "distribution",
                                     "delivery", "hauling", "haul", "van lines"]),
]

# --- Scoring thresholds ---------------------------------------------------------
PRIVATE_SCORE_UNCLEAR_LOW = 35    # below/at -> "ownership unclear" review tab
PRIVATE_SCORE_UNCLEAR_HIGH = 55   # between LOW and HIGH w/ corporate signals -> unclear
SELF_SERVICE_TAG_THRESHOLD = 62   # >= -> RUNS_OWN_SHOP tag

# --- FMCSA Socrata datasets (data.transportation.gov) ---------------------------
SOCRATA_DOMAIN = "https://data.transportation.gov"
DS_CENSUS = "az4n-8mr2"   # Company Census
DS_SMS = "kjg3-diqy"      # SMS Census (dba, passenger/govt flags)
DS_INSPECT = "fx4q-ay7w"  # Inspections (OOS aggregate per carrier)

# --- Apify actors (PINNED). Prices are fetched live at runtime; these are the
#     documented fallbacks (USD per result) if the API price read fails. ---------
APIFY_ACTORS = {
    "maps":    {"id": "2Mdma1N6Fd0y3QEjR",         # Google Maps Extractor (Oxley's pinned actor)
                "slug": "compass~google-maps-extractor",
                "fallback_price": 0.007, "unit": "place"},
    "contact": {"id": "vdrmota~contact-info-scraper",  # emails / socials from websites
                "slug": "vdrmota~contact-info-scraper",
                "fallback_price": 0.005, "unit": "page"},
    "indeed":  {"id": "misceres~indeed-scraper",       # Indeed job postings
                "slug": "misceres~indeed-scraper",
                "fallback_price": 0.004, "unit": "posting"},
}

# =============================================================================
# HTTP helpers
# =============================================================================

UA = "OxleyFleetIntel/1.0 (+moxley@oxleytireinc.com)"
_session = requests.Session()
_session.headers.update({"User-Agent": UA})


def _socrata_headers():
    h = {"Accept": "application/json"}
    tok = os.environ.get("SOCRATA_APP_TOKEN")
    if tok:
        h["X-App-Token"] = tok
    return h


def http_get(url, params=None, headers=None, timeout=90, retries=5):
    """GET with exponential backoff on network / 5xx / 429."""
    delay = 2.0
    last = None
    for attempt in range(retries):
        try:
            r = _session.get(url, params=params, headers=headers, timeout=timeout)
            if r.status_code in (429, 500, 502, 503, 504):
                raise requests.HTTPError(f"{r.status_code}", response=r)
            r.raise_for_status()
            return r
        except Exception as e:  # noqa: BLE001
            last = e
            if attempt == retries - 1:
                break
            time.sleep(delay)
            delay *= 2
    raise RuntimeError(f"GET failed after {retries} tries: {url} :: {last}")


def socrata_getall(dataset, select=None, where=None, group=None, order=None,
                   page=50000, label="", max_rows=None):
    """Page through a Socrata dataset, returning a list of row dicts.
    If max_rows is set, stop once that many rows are collected."""
    base = f"{SOCRATA_DOMAIN}/resource/{dataset}.json"
    out = []
    offset = 0
    while True:
        params = {"$limit": page, "$offset": offset}
        if select:
            params["$select"] = select
        if where:
            params["$where"] = where
        if group:
            params["$group"] = group
        if order:
            params["$order"] = order
        elif not group:
            params["$order"] = ":id"  # stable pagination
        r = http_get(base, params=params, headers=_socrata_headers())
        rows = r.json()
        out.extend(rows)
        got = len(rows)
        if label:
            print(f"    [{label}] +{got:,} (total {len(out):,})", flush=True)
        if max_rows and len(out) >= max_rows:
            out = out[:max_rows]
            break
        if got < page:
            break
        offset += page
    return out


# =============================================================================
# Layer 1 — FMCSA spine
# =============================================================================

CENSUS_COLS = ("dot_number,legal_name,phy_street,phy_city,phy_state,phy_zip,"
               "phy_cnty,phone,email_address,power_units,truck_units,total_drivers,"
               "total_cdl,mcs150_mileage,mcs150_mileage_year,business_org_desc,"
               "company_officer_1,company_officer_2,carrier_operation,hm_ind,"
               "status_code,owntract,owntrail,add_date")

SMS_COLS = ("dot_number,dba_name,pc_flag,private_passenger_business,"
            "private_passenger_nonbusiness,federal_government,state_government,"
            "local_government,indian_tribe,us_mail,authorized_for_hire,"
            "private_property,private_only,hm_flag,nbr_power_unit,driver_total,"
            "recent_mileage")


def pull_fmcsa(states, per_state_limit=None):
    """Pull census + SMS + inspection-OOS aggregate for the given states; join by DOT."""
    carriers = {}  # dot_number -> merged dict

    for st in states:
        print(f"  Census {st} (status A) ...", flush=True)
        where = f"phy_state='{st}' AND status_code='A'"
        # per_state_limit only caps the census spine (rows we actually keep);
        # it must NOT shrink the page size, or the SMS/inspection full pulls
        # below would fan out into hundreds of tiny requests.
        rows = socrata_getall(DS_CENSUS, select=CENSUS_COLS, where=where,
                              page=(min(per_state_limit, 50000) if per_state_limit
                                    else 50000),
                              label=f"census {st}", max_rows=per_state_limit)
        for row in rows:
            dot = str(row.get("dot_number", "")).strip()
            if not dot:
                continue
            carriers[dot] = dict(row)

    # SMS enrichment (dba + passenger/govt flags)
    for st in states:
        print(f"  SMS census {st} ...", flush=True)
        where = f"phy_state='{st}'"
        rows = socrata_getall(DS_SMS, select=SMS_COLS, where=where,
                              page=50000, label=f"sms {st}")
        for row in rows:
            dot = str(row.get("dot_number", "")).strip()
            c = carriers.get(dot)
            if c:
                for k, v in row.items():
                    if k != "dot_number":
                        c.setdefault(k, v)

    # Inspection OOS aggregate per carrier (single grouped query per state)
    for st in states:
        print(f"  Inspection OOS aggregate {st} ...", flush=True)
        # OOS/viol columns are stored as TEXT in this dataset -> cast to number.
        sel = ("dot_number,count(1) as insp_cnt,"
               "sum(vehicle_oos_total::number) as veh_oos,"
               "sum(vehicle_viol_total::number) as veh_viol,"
               "sum(oos_total::number) as oos_all")
        rows = socrata_getall(DS_INSPECT, select=sel,
                              where=f"report_state='{st}'", group="dot_number",
                              order="dot_number", label=f"insp {st}")
        for row in rows:
            dot = str(row.get("dot_number", "")).strip()
            c = carriers.get(dot)
            if c:
                c["insp_cnt"] = _num(row.get("insp_cnt"))
                c["veh_oos"] = _num(row.get("veh_oos"))
                c["veh_viol"] = _num(row.get("veh_viol"))
                c["oos_all"] = _num(row.get("oos_all"))

    return list(carriers.values())


# =============================================================================
# small utils
# =============================================================================

def _num(v, default=0.0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _name_blob(c):
    return f"{c.get('legal_name','')} {c.get('dba_name','')}".lower()


def _truthy_flag(v):
    """Socrata boolean-ish flags arrive as 'Y'/'N'/True/'1'/None."""
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("y", "yes", "true", "1", "t")


def contains_any(text, needles):
    t = text.lower()
    for n in needles:
        if n in t:
            return n
    return None


# =============================================================================
# Layer 2 — screens
# =============================================================================

def is_passenger_or_govt(c):
    """Return a reason string if carrier is passenger/govt/school, else None."""
    for flag in ("pc_flag", "private_passenger_business",
                 "private_passenger_nonbusiness", "federal_government",
                 "state_government", "local_government", "indian_tribe", "us_mail"):
        if _truthy_flag(c.get(flag)):
            return flag
    hit = contains_any(_name_blob(c), PASSENGER_GOVT_NAME)
    if hit:
        return f"name:{hit.strip()}"
    return None


def blocklisted(c):
    return contains_any(_name_blob(c), NATIONAL_BLOCKLIST)


def looks_public(c):
    return contains_any(_name_blob(c), PUBLIC_MARKERS)


def score_private_ownership(c):
    """0-100 soft score + reasons + hard_exclude flag (public company)."""
    score = 40  # neutral prior
    reasons = []
    name = _name_blob(c)
    org = (c.get("business_org_desc") or "").upper()

    # owner surname in company name
    surname_hit = False
    for off in (c.get("company_officer_1"), c.get("company_officer_2")):
        if not off:
            continue
        tokens = re.split(r"[,\s]+", str(off).strip())
        for tok in tokens:
            tok = tok.strip().lower()
            if len(tok) >= 3 and tok not in ("llc", "inc", "the") and tok in name:
                surname_hit = True
                break
        if surname_hit:
            break
    if surname_hit:
        score += 25
        reasons.append("+25 owner name in company name")

    # legal structure
    if org in ("INDIVIDUAL", "PARTNERSHIP"):
        score += 20
        reasons.append(f"+20 org={org.lower()}")
    if " llc" in name or "l.l.c" in name:
        score += 10
        reasons.append("+10 LLC")

    # registered-agent / officer looks like a corporate-services firm -> penalize
    agent_hit = contains_any(name, CORPORATE_AGENTS) or contains_any(
        f"{c.get('company_officer_1','')} {c.get('company_officer_2','')}".lower(),
        CORPORATE_AGENTS)
    if agent_hit:
        score -= 20
        reasons.append(f"-20 corporate-services agent ({agent_hit.strip()})")
    elif c.get("company_officer_1"):
        score += 10
        reasons.append("+10 individual officer on file")

    # public-company signal -> hard exclude
    hard_exclude = False
    pub = looks_public(c)
    if pub:
        hard_exclude = True
        score = 0
        reasons.append(f"HARD-EXCLUDE public marker '{pub.strip()}'")

    # very large corporate fleet w/ no owner-name signal -> likely PE/corporate parent
    if org == "CORPORATION" and _num(c.get("power_units")) >= 250 and not surname_hit:
        score -= 15
        reasons.append("-15 large corporation, no owner-name (possible parent)")

    score = max(0, min(100, score))
    return score, reasons, hard_exclude


# =============================================================================
# Layer 3 — verticals + tire-burn tier
# =============================================================================

def classify_vertical(c):
    blob = _name_blob(c)
    # tire-burn tier
    if contains_any(blob, SEVERE_KEYWORDS):
        tier = "Severe"
    elif contains_any(blob, HEAVY_KEYWORDS):
        tier = "Heavy"
    else:
        tier = "General"
    # vertical label
    for label, kws in VERTICAL_RULES:
        if contains_any(blob, kws):
            return label, tier
    return "Other / Unclassified", tier


# =============================================================================
# Layer 4 — self-service (runs-own-shop) score, FMCSA-only
# =============================================================================

def score_self_service(c):
    """0-100. High power/driver ratio + low vehicle-OOS rate + owns equipment +
    mid-size fleet + real utilization => likely runs its own maintenance shop."""
    score = 0.0
    reasons = []

    pu = _num(c.get("power_units")) or _num(c.get("nbr_power_unit"))
    drv = _num(c.get("total_drivers")) or _num(c.get("driver_total"))
    miles = _num(c.get("mcs150_mileage")) or _num(c.get("recent_mileage"))

    # power-units / drivers ratio  (>=1 == trucks waiting; strong own-shop signal)
    if drv > 0 and pu > 0:
        ratio = pu / drv
        comp = max(0.0, min(1.0, (ratio - 0.6) / 0.9)) * 30  # 0.6->0, 1.5->30
        score += comp
        if comp > 0:
            reasons.append(f"+{comp:.0f} power/driver ratio {ratio:.2f}")

    # vehicle OOS rate vs inspections (lower == better maintained == own shop)
    insp = _num(c.get("insp_cnt"))
    if insp >= 3:
        oos_rate = _num(c.get("veh_oos")) / insp
        comp = max(0.0, min(1.0, (0.35 - oos_rate) / 0.35)) * 30  # 0% oos->30, 35%+->0
        score += comp
        reasons.append(f"+{comp:.0f} veh-OOS rate {oos_rate:.2f} over {int(insp)} insp")
    else:
        score += 8  # unknown -> small neutral credit
        reasons.append("+8 thin inspection history (neutral)")

    # owns tractors / trailers (not leased)
    if _truthy_flag(c.get("owntract")) or _num(c.get("owntract")) > 0:
        score += 8
        reasons.append("+8 owns tractors")
    if _truthy_flag(c.get("owntrail")) or _num(c.get("owntrail")) > 0:
        score += 4
        reasons.append("+4 owns trailers")

    # mid-size fleet band most likely to justify an in-house shop
    if 8 <= pu <= 175:
        score += 12
        reasons.append(f"+12 fleet size {int(pu)} in own-shop band")
    elif pu >= 176:
        score += 6
        reasons.append(f"+6 large fleet {int(pu)}")

    # real utilization (miles per truck)
    if pu > 0 and miles > 0:
        mpt = miles / pu
        if mpt >= 40000:
            score += 6
            reasons.append(f"+6 {int(mpt):,} mi/truck utilization")

    score = max(0, min(100, round(score)))
    return score, reasons


# =============================================================================
# Layer 5 — geocode + residential split
# =============================================================================

PO_BOX_RE = re.compile(r"\b(p\.?\s*o\.?\s*box|post\s+office\s+box|hc\s+\d|rr\s+\d|"
                       r"rural\s+route|route\s+\d|general\s+delivery)\b", re.I)
RESIDENTIAL_HINT_RE = re.compile(r"\b(apt|apartment|unit|lot|trlr|trailer|#)\b", re.I)


def quick_noncommercial_flag(c):
    """Cheap, geocode-free heuristic used for the count preview and as a prefilter.
    Returns (is_noncommercial, reason)."""
    street = (c.get("phy_street") or "")
    if PO_BOX_RE.search(street):
        return True, "po-box/rural-route"
    pu = _num(c.get("power_units"))
    org = (c.get("business_org_desc") or "").upper()
    if RESIDENTIAL_HINT_RE.search(street) and pu <= 3:
        return True, "residential-street-marker, small fleet"
    if org == "INDIVIDUAL" and pu <= 2 and not street.strip():
        return True, "individual, no street, tiny fleet"
    return False, ""


def census_geocode(carriers, batch=5000):
    """Free Census batch geocoder. Adds lat/lng/match/geo_commercial to carriers
    in place. Best-effort: residential vs commercial from PO-box + match quality.
    Only call on a bounded candidate set (geocoding 200k is slow)."""
    url = "https://geocoding.geo.census.gov/geocoder/locations/addressbatch"
    idx = {}
    for i, c in enumerate(carriers):
        idx[str(i)] = c
    items = list(idx.items())
    for start in range(0, len(items), batch):
        chunk = items[start:start + batch]
        buf = io.StringIO()
        w = csv.writer(buf)
        for rid, c in chunk:
            w.writerow([rid, c.get("phy_street", ""), c.get("phy_city", ""),
                        c.get("phy_state", ""), c.get("phy_zip", "")])
        files = {"addressFile": ("addr.csv", buf.getvalue(), "text/csv")}
        data = {"benchmark": "Public_AR_Current"}
        try:
            r = _session.post(url, files=files, data=data, timeout=300)
            r.raise_for_status()
        except Exception as e:  # noqa: BLE001
            print(f"    geocode batch failed ({e}); continuing", flush=True)
            continue
        rdr = csv.reader(io.StringIO(r.text))
        for parts in rdr:
            if not parts:
                continue
            rid = parts[0]
            c = idx.get(rid)
            if not c:
                continue
            match = parts[2] if len(parts) > 2 else "No_Match"
            c["geo_match"] = match
            if match == "Match" and len(parts) >= 6:
                coords = parts[5].split(",") if parts[5] else []
                if len(coords) == 2:
                    c["lng"], c["lat"] = coords[0], coords[1]
            # commercial vs residential: PO box/rural => residential; else assume
            # commercial when it geocodes cleanly to a street address.
            nc, reason = quick_noncommercial_flag(c)
            c["geo_commercial"] = not nc and match == "Match"
            c["geo_noncommercial_reason"] = reason
        time.sleep(0.5)


# =============================================================================
# Region bucketing
# =============================================================================

def assign_region(c):
    st = (c.get("phy_state") or "").upper()
    zip5 = (c.get("phy_zip") or "")[:5]
    if st == "LA":
        return "Louisiana"
    if zip5 in SETX_ZIPS:
        return "SETX"
    z3 = zip5[:3]
    return REGION_ZIP3.get(z3, "rest-of-TX")


# =============================================================================
# Enrichment orchestration (free layers 1-5)
# =============================================================================

def build_free_layer(states, per_state_limit=None, geocode=False):
    print("[1] Pulling FMCSA spine (census + SMS + inspection OOS) ...", flush=True)
    carriers = pull_fmcsa(states, per_state_limit=per_state_limit)
    print(f"    raw active carriers: {len(carriers):,}", flush=True)

    kept, dropped_block, dropped_pax = [], 0, 0
    print("[2-4] Screening + scoring ...", flush=True)
    for c in carriers:
        if blocklisted(c):
            dropped_block += 1
            continue
        pax = is_passenger_or_govt(c)
        if pax:
            dropped_pax += 1
            continue

        ps, preasons, hard = score_private_ownership(c)
        c["private_score"] = ps
        c["private_reasons"] = "; ".join(preasons)
        c["public_hard_exclude"] = hard

        vertical, tier = classify_vertical(c)
        c["vertical"] = vertical
        c["tire_burn"] = tier

        ss, sreasons = score_self_service(c)
        c["self_service_score"] = ss
        c["self_service_reasons"] = "; ".join(sreasons)
        c["runs_own_shop"] = ss >= SELF_SERVICE_TAG_THRESHOLD

        nc, ncr = quick_noncommercial_flag(c)
        c["noncommercial"] = nc
        c["noncommercial_reason"] = ncr

        c["region"] = assign_region(c)

        # ownership-unclear routing
        if hard:
            c["ownership_bucket"] = "public-excluded"
        elif ps <= PRIVATE_SCORE_UNCLEAR_LOW:
            c["ownership_bucket"] = "unclear"
        elif ps <= PRIVATE_SCORE_UNCLEAR_HIGH and contains_any(
                _name_blob(c), CORPORATE_AGENTS):
            c["ownership_bucket"] = "unclear"
        else:
            c["ownership_bucket"] = "private"

        kept.append(c)

    print(f"    dropped (national blocklist): {dropped_block:,}", flush=True)
    print(f"    dropped (passenger/govt/school): {dropped_pax:,}", flush=True)
    print(f"    kept for scoring: {len(kept):,}", flush=True)

    if geocode:
        cand = [c for c in kept if c["region"] == "SETX" or c.get("runs_own_shop")]
        print(f"[5] Geocoding {len(cand):,} candidates (Census batch) ...", flush=True)
        census_geocode(cand)

    return kept


# =============================================================================
# Harvest candidate selection + cost estimate (layer 6, gated)
# =============================================================================

def select_harvest_candidates(carriers, top_n=HARVEST_TOP_N, include_setx=True):
    pool = [c for c in carriers
            if not c["public_hard_exclude"] and c["ownership_bucket"] != "unclear"]
    top = sorted(pool, key=lambda c: c["private_score"], reverse=True)[:top_n]
    chosen = {id(c): c for c in top}
    if include_setx:
        for c in carriers:
            if c["region"] == "SETX" and not c["public_hard_exclude"]:
                chosen[id(c)] = c
    return list(chosen.values())


def fetch_apify_price(actor, token):
    """Best-effort read of the actor's real per-result price from the Apify API.
    Returns (price_usd, source_str). Falls back to the documented price."""
    fb = actor["fallback_price"]
    if not token:
        return fb, "fallback (no APIFY_TOKEN)"
    try:
        url = f"https://api.apify.com/v2/acts/{actor['id']}"
        r = _session.get(url, params={"token": token}, timeout=30)
        r.raise_for_status()
        data = r.json().get("data", {})
        pricing = data.get("pricingInfos") or []
        if pricing:
            latest = pricing[-1]
            ppr = latest.get("pricePerUnitUsd")
            if ppr:
                return float(ppr), "apify-api"
            tiers = latest.get("tieredPricing") or latest.get("pricingPerEvent")
            if tiers:
                return fb, "apify-api (event-priced; using fallback)"
        return fb, "fallback (no per-unit price published)"
    except Exception as e:  # noqa: BLE001
        return fb, f"fallback (api err: {e})"


def estimate_harvest_cost(candidates, token):
    """Print and return the estimated Apify cost for the candidate set."""
    n = len(candidates)
    lines = []
    total = 0.0
    # expected results per fleet per actor
    plan = {
        "maps":    ("Google Maps place lookup", 1),
        "contact": ("Website email/contact crawl", 1),
        "indeed":  ("Indeed job postings", INDEED_MAX_POSTINGS_PER_FLEET),
    }
    print("\n" + "=" * 68)
    print("APIFY DEEP-HARVEST — COST ESTIMATE (nothing has been spent)")
    print("=" * 68)
    print(f"Candidates to harvest: {n:,}  "
          f"(top-{HARVEST_TOP_N} private-scored statewide + all SETX)")
    print(f"{'Actor':<28}{'unit$':>10}{'src':>8}  {'res/fleet':>9}{'subtotal':>11}")
    print("-" * 68)
    for key, actor in APIFY_ACTORS.items():
        price, src = fetch_apify_price(actor, token)
        label, per = plan[key]
        subtotal = price * per * n
        total += subtotal
        srctag = "api" if src == "apify-api" else "fb"
        print(f"{label:<28}{price:>10.4f}{srctag:>8}  {per:>9}{subtotal:>11.2f}")
        lines.append((label, price, src, per, subtotal))
    print("-" * 68)
    print(f"{'ESTIMATED TOTAL':<46}{'$':>10}{total:>11.2f}")
    print(f"Hard cap (--max-cost / MAX_COST_USD): ${MAX_COST_USD:.2f}")
    print("=" * 68)
    if total > MAX_COST_USD:
        print(f"!! Estimate ${total:.2f} EXCEEDS cap ${MAX_COST_USD:.2f} — harvest "
              f"would ABORT. Lower HARVEST_TOP_N or raise --max-cost.")
    else:
        print(f"Estimate is within cap. Re-run with --harvest --confirm to spend.")
    print("=" * 68 + "\n", flush=True)
    return total, lines


# =============================================================================
# Layer 6 — Apify harvest (only runs when explicitly confirmed)
# =============================================================================

def run_apify_actor(actor_id, run_input, token, timeout=1800):
    """Run an actor synchronously and return dataset items."""
    url = f"https://api.apify.com/v2/acts/{actor_id}/run-sync-get-dataset-items"
    r = _session.post(url, params={"token": token}, json=run_input, timeout=timeout)
    r.raise_for_status()
    return r.json()


def harvest(candidates, token, max_cost):
    """Execute the paid harvest. Assumes confirmation already happened."""
    if not token:
        raise SystemExit("APIFY_TOKEN not set — cannot harvest.")
    est, _ = estimate_harvest_cost(candidates, token)
    if est > max_cost:
        raise SystemExit(f"ABORT: estimate ${est:.2f} exceeds --max-cost ${max_cost:.2f}")
    print(f"Harvesting {len(candidates)} fleets via Apify ...", flush=True)
    for c in candidates:
        name = c.get("legal_name", "")
        city = c.get("phy_city", "")
        query = f"{name} {city} {c.get('phy_state','')}".strip()
        # 1) Google Maps
        try:
            places = run_apify_actor(APIFY_ACTORS["maps"]["id"], {
                "searchStringsArray": [query], "maxCrawledPlacesPerSearch": 1,
                "maxImages": 0, "maxReviews": 0, "language": "en"}, token)
            if places:
                p = places[0]
                c["harvest_website"] = p.get("website")
                c["harvest_maps_phone"] = p.get("phone")
                c["harvest_maps_url"] = p.get("url")
                c["harvest_categories"] = ", ".join(p.get("categories", []) or [])
        except Exception as e:  # noqa: BLE001
            c["harvest_err_maps"] = str(e)
        # 2) contact/email crawl on the discovered website
        site = c.get("harvest_website")
        if site:
            try:
                cinfo = run_apify_actor(APIFY_ACTORS["contact"]["id"], {
                    "startUrls": [{"url": site}], "maxDepth": 1,
                    "maxRequestsPerStartUrl": 3}, token)
                emails, phones = [], []
                for row in cinfo:
                    emails += row.get("emails", []) or []
                    phones += row.get("phones", []) or []
                    text = (row.get("text") or "").lower()
                    if any(k in text for k in ("our own", "in-house", "our shop",
                                               "we service our", "own maintenance",
                                               "own equipment", "our fleet")):
                        c["site_own_shop_language"] = True
                        c["self_service_score"] = min(100, c["self_service_score"] + 12)
                c["harvest_emails"] = ", ".join(sorted(set(emails)))
                c["harvest_phones2"] = ", ".join(sorted(set(phones)))
            except Exception as e:  # noqa: BLE001
                c["harvest_err_contact"] = str(e)
        # 3) Indeed job postings (mechanic / tire tech => strong own-shop signal)
        try:
            jobs = run_apify_actor(APIFY_ACTORS["indeed"]["id"], {
                "position": "diesel mechanic", "location": f"{city}, {c.get('phy_state','')}",
                "company": name, "maxItems": INDEED_MAX_POSTINGS_PER_FLEET}, token)
            hits = [j.get("positionName", "") for j in jobs
                    if any(k in (j.get("positionName", "") + j.get("description", "")).lower()
                           for k in ("mechanic", "tire tech", "tire technician",
                                     "diesel tech", "shop foreman", "fleet maint"))]
            if hits:
                c["mechanic_job_postings"] = "; ".join(hits[:INDEED_MAX_POSTINGS_PER_FLEET])
                c["self_service_score"] = min(100, c["self_service_score"] + 10)
        except Exception as e:  # noqa: BLE001
            c["harvest_err_indeed"] = str(e)
        c["runs_own_shop"] = c["self_service_score"] >= SELF_SERVICE_TAG_THRESHOLD
        time.sleep(0.2)
    print("Harvest complete.", flush=True)


# =============================================================================
# Layer 7 — output (xlsx per region) + Drive upload
# =============================================================================

OUT_COLS = [
    ("dot_number", "DOT#"), ("legal_name", "Legal name"), ("dba_name", "DBA"),
    ("phy_street", "Street"), ("phy_city", "City"), ("phy_state", "ST"),
    ("phy_zip", "ZIP"), ("phy_cnty", "County"), ("phone", "Phone"),
    ("email_address", "Email (FMCSA)"), ("harvest_emails", "Emails (web)"),
    ("harvest_website", "Website"), ("power_units", "Power units"),
    ("total_drivers", "Drivers"), ("mcs150_mileage", "Annual miles"),
    ("business_org_desc", "Org type"), ("company_officer_1", "Officer"),
    ("vertical", "Vertical"), ("tire_burn", "Tire-burn"),
    ("private_score", "Private score"), ("self_service_score", "Own-shop score"),
    ("runs_own_shop", "RUNS OWN SHOP"), ("mechanic_job_postings", "Mechanic jobs"),
    ("site_own_shop_language", "Site own-shop lang"),
    ("private_reasons", "Ownership reasons"),
    ("self_service_reasons", "Own-shop reasons"),
]


def _write_sheet(ws, rows):
    from openpyxl.utils import get_column_letter
    headers = [h for _, h in OUT_COLS]
    ws.append(headers)
    for c in rows:
        ws.append([_cell(c.get(k)) for k, _ in OUT_COLS])
    # header style + freeze + autofilter
    from openpyxl.styles import Font, PatternFill
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def _cell(v):
    if v is True:
        return "YES"
    if v is False or v is None:
        return ""
    return v


def _safe_sheet_title(t):
    t = re.sub(r"[\\/*?:\[\]]", "-", str(t))
    return t[:31] or "Sheet"


def write_region_workbook(region, rows, out_dir):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    commercial = [c for c in rows if not c.get("noncommercial")]
    noncommercial = [c for c in rows if c.get("noncommercial")]
    unclear = [c for c in rows if c.get("ownership_bucket") == "unclear"]
    own_shop = [c for c in commercial if c.get("runs_own_shop")]

    # vertical sheets (commercial, non-unclear)
    by_vert = defaultdict(list)
    for c in commercial:
        if c.get("ownership_bucket") == "unclear":
            continue
        by_vert[c["vertical"]].append(c)
    for vert in sorted(by_vert, key=lambda v: -len(by_vert[v])):
        vrows = sorted(by_vert[vert], key=lambda c: (-c["self_service_score"],
                                                     -c["private_score"]))
        _write_sheet(wb.create_sheet(_safe_sheet_title(vert)), vrows)

    # special tabs
    if own_shop:
        _write_sheet(wb.create_sheet("RUNS OWN SHOP"),
                     sorted(own_shop, key=lambda c: -c["self_service_score"]))
    if noncommercial:
        _write_sheet(wb.create_sheet(_safe_sheet_title("REVIEW - non-commercial")),
                     noncommercial)
    if unclear:
        _write_sheet(wb.create_sheet(_safe_sheet_title("REVIEW - ownership unclear")),
                     unclear)
    if not wb.sheetnames:
        wb.create_sheet("empty")

    os.makedirs(out_dir, exist_ok=True)
    fname = f"OxleyFleetIntel_{region.replace(' ', '_')}.xlsx"
    path = os.path.join(out_dir, fname)
    wb.save(path)
    return path


def write_all_workbooks(carriers, out_dir):
    by_region = defaultdict(list)
    for c in carriers:
        by_region[c["region"]].append(c)
    paths = {}
    for region in REGION_ORDER:
        rows = by_region.get(region, [])
        if not rows and region != "SETX":
            continue
        paths[region] = write_region_workbook(region, rows, out_dir)
        print(f"    wrote {region}: {len(rows):,} carriers -> {paths[region]}",
              flush=True)
    return paths


# --- Google Drive upload (optional) ---------------------------------------------

def drive_service():
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError:
        raise SystemExit("google-api-python-client not installed; run "
                         "`pip install -r intel/requirements.txt` for --drive")
    info = None
    path = os.environ.get("GDRIVE_SERVICE_ACCOUNT_FILE")
    raw = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")
    if path and os.path.exists(path):
        with open(path) as f:
            info = json.load(f)
    elif raw:
        info = json.loads(raw)
    else:
        raise SystemExit("Set GDRIVE_SERVICE_ACCOUNT_FILE or GDRIVE_SERVICE_ACCOUNT_JSON")
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def drive_upsert_folder(svc, name, parent=None):
    q = (f"name='{name}' and mimeType='application/vnd.google-apps.folder' "
         f"and trashed=false")
    if parent:
        q += f" and '{parent}' in parents"
    res = svc.files().list(q=q, fields="files(id)", spaces="drive").execute()
    files = res.get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent:
        meta["parents"] = [parent]
    folder = svc.files().create(body=meta, fields="id").execute()
    fid = folder["id"]
    # share to the owner so it shows up in their Drive
    try:
        svc.permissions().create(fileId=fid, body={
            "type": "user", "role": "writer", "emailAddress": OWNER_EMAIL},
            sendNotificationEmail=False).execute()
    except Exception as e:  # noqa: BLE001
        print(f"    (could not share folder to {OWNER_EMAIL}: {e})", flush=True)
    return fid


def drive_upload(paths):
    from googleapiclient.http import MediaFileUpload
    svc = drive_service()
    fid = drive_upsert_folder(svc, DRIVE_FOLDER_NAME)
    print(f"    Drive folder '{DRIVE_FOLDER_NAME}' = {fid}", flush=True)
    for region, path in paths.items():
        name = os.path.basename(path)
        q = f"name='{name}' and '{fid}' in parents and trashed=false"
        existing = svc.files().list(q=q, fields="files(id)").execute().get("files", [])
        media = MediaFileUpload(
            path, mimetype="application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet", resumable=True)
        if existing:
            svc.files().update(fileId=existing[0]["id"], media_body=media).execute()
            print(f"    updated {name}", flush=True)
        else:
            svc.files().create(body={"name": name, "parents": [fid]},
                               media_body=media, fields="id").execute()
            print(f"    uploaded {name}", flush=True)
    return fid


# =============================================================================
# Layer 8 — diff / renewal
# =============================================================================

STATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "state")


def _snapshot_path(scope):
    return os.path.join(STATE_DIR, f"last_run_{scope}.json")


def load_snapshot(scope):
    p = _snapshot_path(scope)
    if os.path.exists(p):
        with open(p) as f:
            return set(json.load(f).get("dot_numbers", []))
    return set()


def save_snapshot(scope, carriers):
    os.makedirs(STATE_DIR, exist_ok=True)
    with open(_snapshot_path(scope), "w") as f:
        json.dump({"generated": datetime.now(timezone.utc).isoformat(),
                   "count": len(carriers),
                   "dot_numbers": sorted(str(c["dot_number"]) for c in carriers)}, f)


# =============================================================================
# Reporting
# =============================================================================

def print_counts(carriers, scope_label):
    print("\n" + "#" * 68)
    print(f"FREE-LAYER RESULTS — {scope_label}")
    print("#" * 68)
    print(f"Total screened-in carriers: {len(carriers):,}\n")

    reg = Counter(c["region"] for c in carriers)
    print("By region:")
    for r in REGION_ORDER:
        print(f"   {r:<14} {reg.get(r, 0):>8,}")
    print()

    vert = Counter(c["vertical"] for c in carriers)
    print("By vertical:")
    for v, n in vert.most_common():
        print(f"   {v:<28} {n:>8,}")
    print()

    tier = Counter(c["tire_burn"] for c in carriers)
    print("Tire-burn tier:")
    for t in ("Severe", "Heavy", "General"):
        print(f"   {t:<10} {tier.get(t, 0):>8,}")
    print()

    own = sum(1 for c in carriers if c["runs_own_shop"])
    nonc = sum(1 for c in carriers if c["noncommercial"])
    unclear = sum(1 for c in carriers if c["ownership_bucket"] == "unclear")
    pub = sum(1 for c in carriers if c["public_hard_exclude"])
    print(f"RUNS_OWN_SHOP tagged:        {own:>8,}")
    print(f"Non-commercial (review):     {nonc:>8,}")
    print(f"Ownership unclear (review):  {unclear:>8,}")
    print(f"Public hard-excluded:        {pub:>8,}")
    print()

    ranges = [(80, 101, "80-100"), (60, 80, "60-79"), (40, 60, "40-59"),
              (0, 40, "0-39")]
    print("Private-ownership score distribution:")
    for lo, hi, lbl in ranges:
        n = sum(1 for c in carriers if lo <= c["private_score"] < hi)
        print(f"   {lbl:<8} {n:>8,}")
    print()

    setx = [c for c in carriers if c["region"] == "SETX"]
    print(f"SETX carriers: {len(setx):,}  "
          f"(own-shop: {sum(1 for c in setx if c['runs_own_shop'])})")
    print("#" * 68 + "\n", flush=True)


# =============================================================================
# CLI
# =============================================================================

def main(argv=None):
    ap = argparse.ArgumentParser(description="Oxley TX+LA fleet-intelligence scraper")
    ap.add_argument("--states", default=",".join(STATES),
                    help="comma list of physical states (default TX,LA)")
    ap.add_argument("--limit", type=int, default=None,
                    help="cap carriers per state (smoke testing only)")
    ap.add_argument("--geocode", action="store_true",
                    help="run Census geocoder on SETX + own-shop candidates")
    ap.add_argument("--out", default=os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "out"),
        help="local output dir for xlsx")
    ap.add_argument("--xlsx", action="store_true",
                    help="also write local xlsx workbooks (free layer)")
    ap.add_argument("--drive", action="store_true",
                    help="upload workbooks to Google Drive (needs SA creds)")
    ap.add_argument("--harvest", action="store_true",
                    help="run the PAID Apify deep-harvest (requires --confirm)")
    ap.add_argument("--confirm", action="store_true",
                    help="non-interactive confirmation for --harvest spend")
    ap.add_argument("--max-cost", type=float, default=MAX_COST_USD,
                    help=f"hard USD cap for harvest (default {MAX_COST_USD})")
    ap.add_argument("--diff", action="store_true",
                    help="SETX-only renewal: diff vs last run, harvest only NEW")
    args = ap.parse_args(argv)

    states = [s.strip().upper() for s in args.states.split(",") if s.strip()]

    # --- diff/renewal mode: SETX only ---
    if args.diff:
        scope = "SETX"
        print(f"[diff] SETX renewal run ...", flush=True)
        carriers = build_free_layer(["TX"], per_state_limit=args.limit,
                                    geocode=args.geocode)
        carriers = [c for c in carriers if c["region"] == "SETX"]
        prev = load_snapshot(scope)
        new = [c for c in carriers if str(c["dot_number"]) not in prev]
        for c in new:
            c["new_this_month"] = True
        print(f"[diff] SETX total {len(carriers):,}; NEW since last run: {len(new):,}",
              flush=True)
        token = os.environ.get("APIFY_TOKEN")
        cand = select_harvest_candidates(new, top_n=len(new), include_setx=True)
        est, _ = estimate_harvest_cost(cand, token)
        if args.harvest:
            if not (args.confirm or _interactive_ok(est)):
                print("Harvest not confirmed; skipping paid layer.", flush=True)
            else:
                harvest(cand, token, args.max_cost)
        paths = write_all_workbooks(carriers, args.out)
        # NEW_THIS_MONTH sheet appended into SETX workbook
        _append_new_sheet(paths.get("SETX"), new)
        if args.drive:
            drive_upload(paths)
        save_snapshot(scope, carriers)
        print_counts(carriers, "SETX renewal")
        return 0

    # --- normal run: free layers 1-5 always; harvest/drive only if asked ---
    carriers = build_free_layer(states, per_state_limit=args.limit,
                                geocode=args.geocode)
    print_counts(carriers, f"states={','.join(states)}"
                 + (f" (limit {args.limit}/state)" if args.limit else ""))

    cand = select_harvest_candidates(carriers, top_n=HARVEST_TOP_N,
                                     include_setx=HARVEST_INCLUDE_SETX)
    token = os.environ.get("APIFY_TOKEN")
    est, _ = estimate_harvest_cost(cand, token)

    if args.xlsx or args.drive:
        paths = write_all_workbooks(carriers, args.out)
        if args.drive:
            drive_upload(paths)

    if args.harvest:
        if not (args.confirm or _interactive_ok(est)):
            print("Harvest NOT confirmed — no Apify credits spent.", flush=True)
            return 0
        harvest(cand, token, args.max_cost)
        paths = write_all_workbooks(carriers, args.out)
        if args.drive:
            drive_upload(paths)
    else:
        print(">> Free layer only. No Apify credits spent. Review counts + estimate "
              "above, then re-run with --harvest --confirm to deep-harvest.",
              flush=True)
    return 0


def _interactive_ok(est):
    if not sys.stdin.isatty():
        return False
    try:
        ans = input(f"Spend ~${est:.2f} on Apify harvest? [y/N] ").strip().lower()
    except EOFError:
        return False
    return ans in ("y", "yes")


def _append_new_sheet(path, new_rows):
    if not path or not os.path.exists(path) or not new_rows:
        return
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if "NEW_THIS_MONTH" in wb.sheetnames:
        del wb["NEW_THIS_MONTH"]
    _write_sheet(wb.create_sheet("NEW_THIS_MONTH", 0), new_rows)
    wb.save(path)


if __name__ == "__main__":
    sys.exit(main())
